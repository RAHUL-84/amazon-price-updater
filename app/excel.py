from openpyxl import Workbook, load_workbook
import os

FILE_NAME = "products.xlsx"


def save_product(serial_no, product, mrp, selling_price, discount):

    # Agar file exist nahi karti to create karo
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active

        ws.append([
            "S.No",
            "Product Name",
            "MRP",
            "Selling Price",
            "Discount"
        ])

        wb.save(FILE_NAME)

    # Existing workbook open karo
    wb = load_workbook(FILE_NAME)
    ws = wb.active

    # Data append karo
    ws.append([
        serial_no,
        product,
        mrp,
        selling_price,
        discount
    ])

    # Save aur close
    wb.save(FILE_NAME)
    wb.close()