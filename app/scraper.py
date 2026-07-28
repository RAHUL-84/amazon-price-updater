from playwright.async_api import async_playwright
from app.config import PRODUCTS, AMAZON_URL
from app.excel import save_product
import os
from openpyxl import load_workbook


async def run_scraper():

    playwright = await async_playwright().start()

    browser = await playwright.chromium.launch(
        headless=False
    )

    page = await browser.new_page()

    page.set_default_timeout(10000)

    print("Opening Amazon...")

    serial = 1

    # Existing Excel se next serial number lo
    if os.path.exists("products.xlsx"):
        wb = load_workbook("products.xlsx")
        ws = wb.active
        serial = ws.max_row
        wb.close()

    for product in PRODUCTS:

        try:

            await page.goto(
                AMAZON_URL,
                wait_until="domcontentloaded"
            )

            print(f"\nSearching : {product}")

            await page.fill(
                "#twotabsearchtextbox",
                product
            )

            await page.keyboard.press("Enter")

            await page.wait_for_timeout(3000)

            result_count = await page.locator(
                "div[data-component-type='s-search-result']"
            ).count()

            print("Results :", result_count)

            first = page.locator(
                "a.a-link-normal.s-no-outline"
            ).first

            href = await first.get_attribute("href")

            print("Href :", href)

            if not href:
                print("Product Link Not Found")
                continue

            await page.goto(
                "https://www.amazon.in" + href,
                wait_until="domcontentloaded"
            )

            await page.wait_for_timeout(2000)

            title_element = page.locator(
                "span#productTitle"
            ).first

            if await title_element.count() == 0:
                print("Title Not Found")
                continue

            title = await title_element.text_content()

            if not title:
                print("Title Not Found")
                continue

            title = title.strip()

            try:
                selling_price = await page.locator(
                    ".a-price .a-offscreen"
                ).first.text_content()
            except:
                selling_price = "N/A"

            try:
                mrp = await page.locator(
                    ".a-text-price .a-offscreen"
                ).first.text_content()
            except:
                mrp = "N/A"

            try:
                discount = await page.locator(
                    "span.savingsPercentage"
                ).first.text_content()
            except:
                discount = "N/A"

            save_product(
                serial,
                title,
                mrp,
                selling_price,
                discount
            )

            print(f"{serial} Saved Successfully")

            serial += 1

        except Exception as e:

            print("ERROR :", e)

    await browser.close()

    await playwright.stop()

    return "Completed"